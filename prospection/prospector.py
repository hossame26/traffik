"""
PROSPECTOR V2 - Systeme de prospection multi-sources
Sources: Google Places API + PagesJaunes + Planity
Output: prospects.xlsx (Excel natif, formate pro)

Usage:
  python3 prospector.py marseille coiffeur       # Une ville, une categorie
  python3 prospector.py marseille coiffeur 30    # Avec limite
  python3 prospector.py --ville marseille        # Toutes categories dans une ville
  python3 prospector.py --all                    # Toutes villes x toutes categories
  python3 prospector.py --planity marseille      # Planity seulement
  python3 prospector.py --pagesjaunes nice coiffeur  # PagesJaunes seulement
"""

import sys
import os
import time
import re
import uuid
import json
import requests
from datetime import datetime
from urllib.parse import urlparse, quote_plus
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from config import GOOGLE_PLACES_KEY, HUNTER_API_KEY, VILLES, CATEGORIES

# --- CONFIG ---
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "prospects.xlsx")

HEADERS_BROWSER = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "fr-FR,fr;q=0.9,en;q=0.8",
}

# Colonnes du fichier Excel
COLUMNS = [
    ("Date", 18),
    ("Nom", 30),
    ("Categorie", 18),
    ("Ville", 16),
    ("Adresse", 40),
    ("Telephone", 18),
    ("Email", 28),
    ("Site Web", 35),
    ("Score Site", 14),
    ("Problemes", 40),
    ("Google Maps", 35),
    ("Note Google", 12),
    ("Nb Avis", 10),
    ("Source", 14),
    ("Statut", 14),
    ("Notes", 30),
]

# Score couleurs
SCORE_COLORS = {
    "PAS DE SITE": "FF1744",   # Rouge
    "MAUVAIS": "FF6D00",       # Orange
    "MOYEN": "FFD600",         # Jaune
    "CORRECT": "00C853",       # Vert
    "ERREUR": "9E9E9E",        # Gris
}


# ============================================================
# SOURCES DE DONNEES
# ============================================================

def scrape_google_places(ville, categorie, max_results=20):
    """Google Places Text Search API"""
    if not GOOGLE_PLACES_KEY:
        print("  [!] Pas de cle Google Places API dans .env")
        return []

    url = "https://maps.googleapis.com/maps/api/place/textsearch/json"
    query = f"{categorie} {ville}"
    results = []

    try:
        resp = requests.get(url, params={
            "query": query,
            "key": GOOGLE_PLACES_KEY,
            "language": "fr",
            "region": "fr"
        }, timeout=10)
        data = resp.json()

        if data.get("status") != "OK":
            print(f"  [!] Google Places: {data.get('status')} - {data.get('error_message', '')}")
            return []

        places = data.get("results", [])[:max_results]
        print(f"  [Google] {len(places)} resultats bruts")

        for place in places:
            # Details supplementaires
            details = _google_get_details(place["place_id"])
            time.sleep(0.5)  # Rate limit

            results.append({
                "nom": place.get("name", ""),
                "categorie": categorie,
                "ville": ville,
                "adresse": place.get("formatted_address", ""),
                "telephone": details.get("telephone", ""),
                "site_web": details.get("site_web", ""),
                "google_maps": details.get("google_maps", ""),
                "rating": place.get("rating", ""),
                "nb_avis": place.get("user_ratings_total", ""),
                "source": "Google Places",
            })

    except Exception as e:
        print(f"  [!] Google Places erreur: {e}")

    return results


def _google_get_details(place_id):
    """Details d'un commerce via Google Places Details API"""
    try:
        resp = requests.get(
            "https://maps.googleapis.com/maps/api/place/details/json",
            params={
                "place_id": place_id,
                "key": GOOGLE_PLACES_KEY,
                "fields": "formatted_phone_number,website,url",
                "language": "fr"
            },
            timeout=8
        )
        result = resp.json().get("result", {})
        return {
            "telephone": result.get("formatted_phone_number", ""),
            "site_web": result.get("website", ""),
            "google_maps": result.get("url", ""),
        }
    except Exception:
        return {}


def scrape_pagesjaunes(ville, categorie, max_results=20):
    """Scrape PagesJaunes.fr via Playwright (Cloudflare bypass)"""
    results = []
    url = f"https://www.pagesjaunes.fr/annuaire/chercherlespros?quoiqui={quote_plus(categorie)}&ou={quote_plus(ville)}"

    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx = browser.new_context(
                viewport={"width": 1920, "height": 1080},
                user_agent=HEADERS_BROWSER["User-Agent"],
                locale="fr-FR",
            )
            page = ctx.new_page()

            # Homepage d'abord (cookies Cloudflare)
            page.goto("https://www.pagesjaunes.fr", wait_until="domcontentloaded", timeout=15000)
            page.wait_for_timeout(1500)

            # Accepter cookies
            try:
                page.click("#didomi-notice-agree-button", timeout=3000)
                page.wait_for_timeout(500)
            except Exception:
                pass

            # Recherche
            page.goto(url, wait_until="domcontentloaded", timeout=20000)
            page.wait_for_timeout(4000)

            # Parser le HTML
            html = page.content()
            browser.close()

        soup = BeautifulSoup(html, "html.parser")
        cards = soup.select("li.bi")
        print(f"  [PagesJaunes] {len(cards)} resultats")

        for card in cards[:max_results]:
            try:
                # Nom
                name_el = card.select_one("a.bi-denomination")
                nom = name_el.get_text(strip=True) if name_el else ""
                if not nom:
                    continue

                # Adresse
                addr_el = card.select_one(".bi-address a.pj-link")
                adresse_raw = addr_el.get_text(strip=True) if addr_el else ""
                adresse = adresse_raw.replace("Voir le plan", "").strip()

                # Telephone
                phone_el = card.select_one(".bi-fantomas .number-contact, .number-contact")
                phone_raw = phone_el.get_text(strip=True) if phone_el else ""
                phone_match = re.search(r"(\d[\d\s]{8,})", phone_raw)
                telephone = phone_match.group(1).strip() if phone_match else ""

                # Note Google
                google_el = card.select_one(".pjts_avis-partenaire strong.txt_l")
                rating = google_el.get_text(strip=True) if google_el else ""

                # Nb avis Google
                google_avis_el = card.select_one(".pjts_avis-partenaire .txt_s")
                nb_avis = ""
                if google_avis_el:
                    m = re.search(r"(\d+)", google_avis_el.get_text())
                    nb_avis = m.group(1) if m else ""

                # Site web
                site_el = card.select_one('a[data-pjlb*="site_internet"]')
                site_web = site_el.get("href", "") if site_el else ""
                if "pagesjaunes.fr" in site_web:
                    site_web = ""

                results.append({
                    "nom": nom,
                    "categorie": categorie,
                    "ville": ville,
                    "adresse": adresse,
                    "telephone": telephone,
                    "site_web": site_web,
                    "google_maps": "",
                    "rating": rating,
                    "nb_avis": nb_avis,
                    "source": "PagesJaunes",
                })
            except Exception:
                continue

    except ImportError:
        print("  [!] Playwright non installe. pip install playwright")
    except Exception as e:
        print(f"  [!] PagesJaunes erreur: {e}")

    return results


def scrape_planity(ville, max_results=20):
    """Scrape Planity.com via Playwright (React SSR)"""
    results = []
    search_url = f"https://www.planity.com/recherche?q=&loc={quote_plus(ville)}"

    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            ctx = browser.new_context(
                viewport={"width": 1920, "height": 1080},
                user_agent=HEADERS_BROWSER["User-Agent"],
                locale="fr-FR",
            )
            page = ctx.new_page()
            page.goto(search_url, wait_until="domcontentloaded", timeout=20000)
            page.wait_for_timeout(5000)

            # Accepter cookies
            try:
                page.click("button[id*='accept'], [class*='cookie'] button", timeout=3000)
                page.wait_for_timeout(500)
            except Exception:
                pass

            # Scroll pour charger plus de resultats
            for _ in range(3):
                page.evaluate("window.scrollBy(0, 800)")
                page.wait_for_timeout(1000)

            html = page.content()
            browser.close()

        soup = BeautifulSoup(html, "html.parser")

        # Methode 1: JSON __NEXT_DATA__
        script_tag = soup.find("script", id="__NEXT_DATA__")
        if script_tag:
            try:
                data = json.loads(script_tag.string)
                props = data.get("props", {}).get("pageProps", {})
                # Chercher dans differentes cles possibles
                places = (props.get("places") or props.get("results") or
                          props.get("establishments") or props.get("searchResults") or [])

                if isinstance(places, dict):
                    places = list(places.values())

                print(f"  [Planity] {len(places)} resultats (JSON)")

                for place in places[:max_results]:
                    if isinstance(place, str):
                        continue
                    nom = place.get("name", "") or place.get("title", "")
                    if not nom:
                        continue

                    addr = place.get("address", {})
                    if isinstance(addr, str):
                        adresse = addr
                    else:
                        adresse = f"{addr.get('street', '')} {addr.get('zipCode', '')} {addr.get('city', '')}".strip()

                    results.append({
                        "nom": nom,
                        "categorie": place.get("category", "beaute"),
                        "ville": ville,
                        "adresse": adresse,
                        "telephone": place.get("phone", "") or place.get("phoneNumber", ""),
                        "site_web": place.get("website", "") or place.get("websiteUrl", ""),
                        "google_maps": "",
                        "rating": place.get("rating", "") or place.get("averageRating", ""),
                        "nb_avis": place.get("reviewCount", "") or place.get("reviewsCount", ""),
                        "source": "Planity",
                    })
            except (json.JSONDecodeError, KeyError, TypeError):
                pass

        # Methode 2: parser le HTML rendu
        if not results:
            # Planity utilise des classes CSS generees, chercher les patterns
            all_links = soup.find_all("a", href=re.compile(r"/[a-z]+-[a-z]+-\d+"))
            seen_names = set()

            for link in all_links[:max_results * 2]:
                parent = link.find_parent(["div", "li", "article"])
                if not parent:
                    continue

                text = link.get_text(strip=True)
                if not text or len(text) < 3 or text in seen_names:
                    continue
                seen_names.add(text)

                # Chercher adresse dans le parent
                addr_text = ""
                for child in parent.find_all(text=re.compile(r"\d{5}")):
                    addr_text = child.strip()
                    break

                results.append({
                    "nom": text,
                    "categorie": "beaute",
                    "ville": ville,
                    "adresse": addr_text,
                    "telephone": "",
                    "site_web": "",
                    "google_maps": "",
                    "rating": "",
                    "nb_avis": "",
                    "source": "Planity",
                })

            if results:
                print(f"  [Planity] {len(results)} resultats (HTML)")

        if not results:
            print(f"  [Planity] 0 resultats")

    except ImportError:
        print("  [!] Playwright non installe. pip install playwright")
    except Exception as e:
        print(f"  [!] Planity erreur: {e}")

    return results


# ============================================================
# ENRICHISSEMENT
# ============================================================

def check_site_quality(url):
    """Analyse la qualite d'un site web"""
    if not url:
        return "PAS DE SITE", "Aucun site web"

    try:
        resp = requests.get(url, timeout=8, allow_redirects=True, headers=HEADERS_BROWSER)
        load_time = resp.elapsed.total_seconds()
        html = resp.text.lower()
        issues = []

        if not resp.url.startswith("https"):
            issues.append("Pas SSL")
        if load_time > 3:
            issues.append(f"Lent ({load_time:.1f}s)")
        if "viewport" not in html:
            issues.append("Pas responsive")
        if '<meta name="description"' not in html:
            issues.append("Pas meta desc")
        if "<h1" not in html:
            issues.append("Pas H1")
        if not any(x in html for x in ["google-analytics", "gtag", "gtm"]):
            issues.append("Pas Analytics")

        # Checks supplementaires V2
        if "wix" in html or "weebly" in html or "jimdo" in html:
            issues.append("Builder basique")
        if resp.status_code >= 400:
            issues.append(f"HTTP {resp.status_code}")

        if len(issues) >= 4:
            score = "MAUVAIS"
        elif len(issues) >= 2:
            score = "MOYEN"
        else:
            score = "CORRECT"

        return score, " | ".join(issues) if issues else "RAS"

    except requests.exceptions.SSLError:
        return "MAUVAIS", "Erreur SSL"
    except requests.exceptions.Timeout:
        return "MAUVAIS", "Timeout (>8s)"
    except requests.exceptions.ConnectionError:
        return "ERREUR", "Site inaccessible"
    except Exception as e:
        return "ERREUR", str(e)[:60]


def find_email(site_url):
    """Trouve l'email d'un commerce"""
    if not site_url:
        return ""

    domain = urlparse(site_url).netloc or site_url
    domain = domain.replace("www.", "")

    # 1. Hunter.io si dispo
    if HUNTER_API_KEY:
        try:
            resp = requests.get(
                "https://api.hunter.io/v2/domain-search",
                params={"domain": domain, "api_key": HUNTER_API_KEY, "limit": 1},
                timeout=8
            )
            data = resp.json()
            emails = data.get("data", {}).get("emails", [])
            if emails:
                return emails[0]["value"]
        except Exception:
            pass

    # 2. Scraper la page contact du site
    try:
        for path in ["", "/contact", "/nous-contacter", "/contactez-nous"]:
            resp = requests.get(f"{site_url.rstrip('/')}{path}", timeout=5, headers=HEADERS_BROWSER)
            emails_found = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', resp.text)
            # Filtrer les emails generiques (pas de noreply, pas de @wix, etc.)
            valid = [e for e in emails_found if not any(x in e.lower() for x in
                     ["noreply", "no-reply", "@wix", "@sentry", "@google", "@facebook",
                      "example.com", "email.com", "domain.com", "@w3.org"])]
            if valid:
                return valid[0]
    except Exception:
        pass

    # 3. Pattern guessing
    return f"contact@{domain}"


def enrich_prospect(prospect):
    """Enrichit un prospect avec score site + email"""
    site = prospect.get("site_web", "")

    # Score du site
    score, problemes = check_site_quality(site)
    prospect["score_site"] = score
    prospect["problemes"] = problemes

    # Email
    prospect["email"] = find_email(site) if site else ""

    return prospect


# ============================================================
# DEDUPLICATION
# ============================================================

def load_existing_prospects():
    """Charge les prospects existants depuis le fichier Excel"""
    existing = set()
    if not os.path.exists(EXCEL_FILE):
        return existing

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[1]:  # Colonne Nom (index 1)
                # Cle de dedup: nom normalise + ville
                key = f"{str(row[1]).lower().strip()}|{str(row[3]).lower().strip() if row[3] else ''}"
                existing.add(key)
        wb.close()
    except Exception:
        pass

    return existing


def is_duplicate(prospect, existing_keys):
    """Verifie si un prospect existe deja"""
    key = f"{prospect['nom'].lower().strip()}|{prospect['ville'].lower().strip()}"
    return key in existing_keys


# ============================================================
# EXPORT EXCEL
# ============================================================

def create_excel_if_needed():
    """Cree le fichier Excel avec en-tetes et formatage"""
    if os.path.exists(EXCEL_FILE):
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Prospects"

    # En-tetes
    header_font = Font(name="Inter", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Figer la premiere ligne
    ws.freeze_panes = "A2"

    # Filtres auto
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    wb.save(EXCEL_FILE)
    wb.close()


def save_to_excel(prospects):
    """Ajoute les prospects au fichier Excel avec formatage"""
    create_excel_if_needed()

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # Styles
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )
    default_font = Font(name="Inter", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    added = 0
    for p in prospects:
        row_num = ws.max_row + 1
        row_data = [
            datetime.now().strftime("%Y-%m-%d %H:%M"),
            p.get("nom", ""),
            p.get("categorie", ""),
            p.get("ville", ""),
            p.get("adresse", ""),
            p.get("telephone", ""),
            p.get("email", ""),
            p.get("site_web", ""),
            p.get("score_site", ""),
            p.get("problemes", ""),
            p.get("google_maps", ""),
            p.get("rating", ""),
            p.get("nb_avis", ""),
            p.get("source", ""),
            "A_CONTACTER",
            "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = default_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx in [1, 3, 4, 9, 12, 13, 14, 15] else left_align

        # Colorer le score
        score = p.get("score_site", "")
        if score in SCORE_COLORS:
            score_cell = ws.cell(row=row_num, column=9)
            score_cell.fill = PatternFill(start_color=SCORE_COLORS[score], end_color=SCORE_COLORS[score], fill_type="solid")
            if score in ("PAS DE SITE", "MAUVAIS"):
                score_cell.font = Font(name="Inter", size=10, bold=True, color="FFFFFF")
            else:
                score_cell.font = Font(name="Inter", size=10, bold=True)

        # Colorer le statut
        statut_cell = ws.cell(row=row_num, column=15)
        statut_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        statut_cell.font = Font(name="Inter", size=10, bold=True, color="1565C0")

        added += 1

    wb.save(EXCEL_FILE)
    wb.close()
    return added


# ============================================================
# MIGRATION CSV -> XLSX
# ============================================================

def migrate_csv_to_excel():
    """Migre l'ancien CSV vers le nouveau format Excel"""
    import csv
    csv_file = os.path.join(os.path.dirname(__file__), "prospects.csv")
    if not os.path.exists(csv_file):
        return 0

    if os.path.exists(EXCEL_FILE):
        # Verifier si deja migre
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row > 1:
            wb.close()
            return 0
        wb.close()

    prospects = []
    with open(csv_file, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            prospects.append({
                "nom": row.get("nom", ""),
                "categorie": row.get("categorie", ""),
                "ville": row.get("ville", ""),
                "adresse": row.get("adresse", ""),
                "telephone": row.get("telephone", ""),
                "email": row.get("email", ""),
                "site_web": row.get("site_web", ""),
                "score_site": row.get("score_site", ""),
                "problemes": row.get("problemes", ""),
                "google_maps": row.get("google_maps", ""),
                "rating": row.get("rating", ""),
                "nb_avis": row.get("nb_avis", ""),
                "source": "Migration CSV",
            })

    if prospects:
        count = save_to_excel(prospects)
        print(f"  [Migration] {count} prospects migres depuis CSV -> Excel")
        return count
    return 0


# ============================================================
# CLI PRINCIPAL
# ============================================================

def print_banner():
    print("\n" + "=" * 60)
    print("  PROSPECTOR V2 - Multi-Sources Lead Generation")
    print("  Sources: Google Places | PagesJaunes | Planity")
    print("=" * 60)


def print_summary(all_prospects, new_count):
    """Affiche le resume final"""
    total = len(all_prospects)
    sans_site = sum(1 for p in all_prospects if p.get("score_site") == "PAS DE SITE")
    mauvais = sum(1 for p in all_prospects if p.get("score_site") == "MAUVAIS")
    moyen = sum(1 for p in all_prospects if p.get("score_site") == "MOYEN")
    correct = sum(1 for p in all_prospects if p.get("score_site") == "CORRECT")
    erreur = sum(1 for p in all_prospects if p.get("score_site") == "ERREUR")
    with_email = sum(1 for p in all_prospects if p.get("email"))

    print(f"\n{'=' * 60}")
    print(f"  RESUME")
    print(f"{'=' * 60}")
    print(f"  Total analyses      : {total}")
    print(f"  Nouveaux ajoutes    : {new_count}")
    print(f"  Avec email          : {with_email}")
    print(f"{'=' * 60}")
    print(f"  SANS SITE (chaud)   : {sans_site}")
    print(f"  MAUVAIS (chaud)     : {mauvais}")
    print(f"  MOYEN (tiede)       : {moyen}")
    print(f"  CORRECT             : {correct}")
    print(f"  ERREUR              : {erreur}")
    print(f"{'=' * 60}")
    print(f"  Fichier: {EXCEL_FILE}")
    print(f"{'=' * 60}\n")

    # Top prospects (sans site ou mauvais)
    hot = [p for p in all_prospects if p.get("score_site") in ("PAS DE SITE", "MAUVAIS")]
    if hot:
        print("  TOP PROSPECTS (a contacter en priorite):")
        for i, p in enumerate(hot[:10], 1):
            phone = f" | {p['telephone']}" if p.get('telephone') else ""
            email = f" | {p['email']}" if p.get('email') else ""
            print(f"    {i}. {p['nom']} ({p['categorie']}, {p['ville']}){phone}{email}")
        print()


def run_prospection(villes, categories, max_per_source=20, sources=None):
    """Lance la prospection sur les villes et categories donnees"""
    if sources is None:
        sources = ["google", "pagesjaunes"]

    all_prospects = []
    existing = load_existing_prospects()
    total_combos = len(villes) * len(categories) if "planity" not in sources else len(villes)

    combo_idx = 0
    for ville in villes:
        # Planity: pas besoin de categorie (beaute seulement)
        if "planity" in sources:
            combo_idx += 1
            print(f"\n[{combo_idx}/{total_combos}] Planity - {ville}")
            raw = scrape_planity(ville, max_per_source)
            for p in raw:
                if not is_duplicate(p, existing):
                    all_prospects.append(p)
                    existing.add(f"{p['nom'].lower().strip()}|{p['ville'].lower().strip()}")

        for categorie in categories:
            combo_idx += 1
            print(f"\n[{combo_idx}/{total_combos}] {ville} > {categorie}")

            raw = []

            if "google" in sources:
                raw += scrape_google_places(ville, categorie, max_per_source)
                time.sleep(1)

            if "pagesjaunes" in sources:
                raw += scrape_pagesjaunes(ville, categorie, max_per_source)
                time.sleep(1)

            # Dedup et enrichissement
            for p in raw:
                if is_duplicate(p, existing):
                    continue

                # Enrichir
                print(f"    -> {p['nom']}...", end=" ", flush=True)
                enrich_prospect(p)

                score_icon = {
                    "PAS DE SITE": "[CHAUD]",
                    "MAUVAIS": "[CHAUD]",
                    "MOYEN": "[TIEDE]",
                    "CORRECT": "[OK]",
                    "ERREUR": "[ERR]",
                }.get(p["score_site"], "?")
                email_info = f" | {p['email']}" if p.get('email') else ""
                print(f"{score_icon} {p['score_site']}{email_info}")

                all_prospects.append(p)
                existing.add(f"{p['nom'].lower().strip()}|{p['ville'].lower().strip()}")

                time.sleep(0.3)

    return all_prospects


def main():
    print_banner()

    # Migrer l'ancien CSV si besoin
    migrate_csv_to_excel()

    args = sys.argv[1:]

    # Modes
    if not args:
        print("\nUsage:")
        print("  python3 prospector.py <ville> <categorie> [max]")
        print("  python3 prospector.py --ville <ville>           # Toutes categories")
        print("  python3 prospector.py --all                     # Tout scraper")
        print("  python3 prospector.py --planity <ville>         # Planity seulement")
        print("  python3 prospector.py --pagesjaunes <ville> <categorie>")
        print(f"\nVilles: {', '.join(VILLES)}")
        print(f"Categories: {', '.join(CATEGORIES[:8])}...")
        sys.exit(0)

    villes = []
    categories_list = []
    max_results = 20
    sources = ["google", "pagesjaunes"]

    if args[0] == "--all":
        villes = VILLES
        categories_list = CATEGORIES
        print(f"\n  Mode: TOUTES villes ({len(villes)}) x categories ({len(categories_list)})")

    elif args[0] == "--ville":
        if len(args) < 2:
            print("[!] Manque la ville. Ex: --ville Marseille")
            sys.exit(1)
        villes = [args[1]]
        categories_list = CATEGORIES
        print(f"\n  Mode: {args[1]} - toutes categories ({len(categories_list)})")

    elif args[0] == "--planity":
        if len(args) < 2:
            print("[!] Manque la ville. Ex: --planity Marseille")
            sys.exit(1)
        villes = [args[1]]
        categories_list = ["beaute"]  # Planity = beaute only
        sources = ["planity"]
        print(f"\n  Mode: Planity - {args[1]}")

    elif args[0] == "--pagesjaunes":
        if len(args) < 3:
            print("[!] Usage: --pagesjaunes <ville> <categorie>")
            sys.exit(1)
        villes = [args[1]]
        categories_list = [args[2]]
        sources = ["pagesjaunes"]
        print(f"\n  Mode: PagesJaunes - {args[1]} > {args[2]}")

    else:
        # Mode simple: ville categorie [max]
        villes = [args[0]]
        categories_list = [args[1]] if len(args) > 1 else CATEGORIES
        max_results = int(args[2]) if len(args) > 2 else 20

    # Lancer
    all_prospects = run_prospection(villes, categories_list, max_results, sources)

    if not all_prospects:
        print("\n  Aucun nouveau prospect trouve.")
        sys.exit(0)

    # Sauvegarder
    new_count = save_to_excel(all_prospects)
    print_summary(all_prospects, new_count)


if __name__ == "__main__":
    main()
